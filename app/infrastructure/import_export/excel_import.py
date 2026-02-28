# app/infrastructure/import_export/excel_import.py
"""
ExcelImportService

Назначение:
- Импорт справочников и учебного плана из Excel (xlsx)
- Запись данных в SQLite (через session_factory, взятый из репозиториев)
- Возврат статистики импорта для GUI

Ожидаемые листы (MVP):
    Teachers
    Subjects
    Groups
    Rooms
    Curriculum
    Calendar
    Availability

Примечание:
- В di.py сервис создаётся с репозиториями (teachers_repo=..., ...)
- Поэтому session_factory извлекаем из репозитория: repo._session_factory
"""

from __future__ import annotations

from pathlib import Path
from typing import Dict, Optional

from openpyxl import load_workbook


class ExcelImportService:
    def __init__(
        self,
        teachers_repo,
        subjects_repo,
        groups_repo,
        rooms_repo,
        calendar_repo,
        curriculum_repo,
        schedule_repo=None,
    ):
        # Репозитории (могут пригодиться позже для более “чистого” импорта)
        self._teachers_repo = teachers_repo
        self._subjects_repo = subjects_repo
        self._groups_repo = groups_repo
        self._rooms_repo = rooms_repo
        self._calendar_repo = calendar_repo
        self._curriculum_repo = curriculum_repo
        self._schedule_repo = schedule_repo

        # Достаём session_factory из любого sqlite-репозитория
        self._session_factory = getattr(teachers_repo, "_session_factory", None)
        if self._session_factory is None:
            raise TypeError(
                "ExcelImportService: не удалось получить session_factory из teachers_repo. "
                "Ожидается Sqlite...Repository с полем _session_factory."
            )

    # ---------------------------------------------------------

    def import_all(self, file_path: Path, fmt: str, flags: Dict[str, bool]) -> Dict:
        if fmt != "xlsx":
            raise ValueError("MVP поддерживает только xlsx")

        wb = load_workbook(file_path)

        stats = {
            "teachers": 0,
            "subjects": 0,
            "groups": 0,
            "rooms": 0,
            "calendar": 0,
            "curriculum": 0,
            "availability": 0,
            "warnings": [],
        }

        with self._session_factory() as conn:
            # Teachers
            if flags.get("teachers") and "Teachers" in wb.sheetnames:
                stats["teachers"] = self._import_teachers(conn, wb["Teachers"])
            elif flags.get("teachers"):
                stats["warnings"].append("Лист Teachers не найден в файле.")

            # Subjects
            if flags.get("subjects") and "Subjects" in wb.sheetnames:
                stats["subjects"] = self._import_subjects(conn, wb["Subjects"])
            elif flags.get("subjects"):
                stats["warnings"].append("Лист Subjects не найден в файле.")

            # Groups
            if flags.get("groups") and "Groups" in wb.sheetnames:
                stats["groups"] = self._import_groups(conn, wb["Groups"])
            elif flags.get("groups"):
                stats["warnings"].append("Лист Groups не найден в файле.")

            # Rooms
            if flags.get("rooms") and "Rooms" in wb.sheetnames:
                stats["rooms"] = self._import_rooms(conn, wb["Rooms"])
            elif flags.get("rooms"):
                stats["warnings"].append("Лист Rooms не найден в файле.")

            # Curriculum
            if flags.get("curriculum") and "Curriculum" in wb.sheetnames:
                stats["curriculum"] = self._import_curriculum(conn, wb["Curriculum"])
            elif flags.get("curriculum"):
                stats["warnings"].append("Лист Curriculum не найден в файле.")

            # Calendar
            if flags.get("calendar") and "Calendar" in wb.sheetnames:
                stats["calendar"] = self._import_calendar(conn, wb["Calendar"])
            elif flags.get("calendar"):
                stats["warnings"].append("Лист Calendar не найден в файле.")

            # Availability
            if flags.get("availability") and "Availability" in wb.sheetnames:
                stats["availability"] = self._import_availability(conn, wb["Availability"])
            elif flags.get("availability"):
                stats["warnings"].append("Лист Availability не найден в файле.")

            conn.commit()

        return stats

    # ---------------------------------------------------------
    # Individual sheet importers
    # ---------------------------------------------------------

    def _import_teachers(self, conn, ws) -> int:
        """
        Ожидаемые колонки:
        A id_teacher
        B full_name
        C max_pairs_per_day_hard (optional)
        D max_pairs_per_day_soft (optional)
        E needs_method_day (optional 0/1)
        """
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue

            id_teacher = int(row[0])
            full_name = str(row[1]).strip() if row[1] is not None else ""

            conn.execute(
                """
                INSERT OR REPLACE INTO Teachers
                (id_teacher, full_name, max_pairs_per_day_hard, max_pairs_per_day_soft, needs_method_day)
                VALUES (?, ?, ?, ?, ?)
                """,
                (
                    id_teacher,
                    full_name,
                    int(row[2]) if row[2] is not None else 6,
                    int(row[3]) if row[3] is not None else 4,
                    int(row[4]) if row[4] is not None else 1,
                ),
            )
            count += 1
        return count

    def _import_subjects(self, conn, ws) -> int:
        """
        A id_subject
        B subject_name
        """
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            conn.execute(
                "INSERT OR REPLACE INTO Subjects(id_subject, subject_name) VALUES (?, ?)",
                (int(row[0]), str(row[1]).strip() if row[1] is not None else ""),
            )
            count += 1
        return count

    def _import_groups(self, conn, ws) -> int:
        """
        A id_group
        B group_name
        C year (optional)
        D quantity
        E education_form (optional)
        """
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            conn.execute(
                """
                INSERT OR REPLACE INTO StudentGroups
                (id_group, group_name, year, quantity, education_form)
                VALUES (?, ?, ?, ?, ?)
                """,
                (
                    int(row[0]),
                    str(row[1]).strip() if row[1] is not None else "",
                    int(row[2]) if row[2] is not None else None,
                    int(row[3]) if row[3] is not None else 0,
                    str(row[4]).strip() if len(row) > 4 and row[4] is not None else "full-time",
                ),
            )
            count += 1
        return count

    def _import_rooms(self, conn, ws) -> int:
        """
        A id_class
        B room_number
        C room_type
        D capacity
        E building (optional)
        """
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            conn.execute(
                """
                INSERT OR REPLACE INTO Classes
                (id_class, room_number, room_type, capacity, building)
                VALUES (?, ?, ?, ?, ?)
                """,
                (
                    int(row[0]),
                    str(row[1]).strip() if row[1] is not None else "",
                    str(row[2]).strip() if row[2] is not None else "",
                    int(row[3]) if row[3] is not None else 0,
                    str(row[4]).strip() if len(row) > 4 and row[4] is not None else None,
                ),
            )
            count += 1
        return count

    def _import_curriculum(self, conn, ws) -> int:
        """
        A id_curriculum
        B group_id
        C subject_id
        D part_type
        E required_room_type
        F hours_total_year (optional)
        """
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            conn.execute(
                """
                INSERT OR REPLACE INTO CurriculumItems
                (id_curriculum, group_id, subject_id, part_type, required_room_type, hours_total_year)
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                (
                    int(row[0]),
                    int(row[1]) if row[1] is not None else None,
                    int(row[2]) if row[2] is not None else None,
                    str(row[3]).strip() if row[3] is not None else "",
                    str(row[4]).strip() if row[4] is not None else "",
                    int(row[5]) if len(row) > 5 and row[5] is not None else 0,
                ),
            )
            count += 1
        return count

    def _import_calendar(self, conn, ws) -> int:
        """
        A id_calendar
        B academic_year
        C semester
        D start_date (optional)
        E end_date (optional)
        F week_type_mode (optional)
        """
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            conn.execute(
                """
                INSERT OR REPLACE INTO AcademicCalendar
                (id_calendar, academic_year, semester, start_date, end_date, week_type_mode)
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                (
                    int(row[0]),
                    str(row[1]).strip() if row[1] is not None else "",
                    int(row[2]) if row[2] is not None else 1,
                    str(row[3]).strip() if len(row) > 3 and row[3] is not None else None,
                    str(row[4]).strip() if len(row) > 4 and row[4] is not None else None,
                    int(row[5]) if len(row) > 5 and row[5] is not None else 1,
                ),
            )
            count += 1
        return count

    def _import_availability(self, conn, ws) -> int:
        """
        A calendar_id
        B teacher_id
        C slot_id
        D is_available (0/1)
        """
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            conn.execute(
                """
                INSERT OR REPLACE INTO TeacherAvailability
                (calendar_id, teacher_id, slot_id, is_available)
                VALUES (?, ?, ?, ?)
                """,
                (
                    int(row[0]),
                    int(row[1]) if row[1] is not None else None,
                    int(row[2]) if row[2] is not None else None,
                    int(row[3]) if row[3] is not None else 1,
                ),
            )
            count += 1
        return count