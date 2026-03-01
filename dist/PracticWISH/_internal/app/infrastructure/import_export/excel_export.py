# app/infrastructure/import_export/excel_export.py
"""
ExcelExportService

Задачи:
- экспорт справочников в Excel/CSV
- экспорт варианта расписания:
    - таблицей (flat list)
    - сеткой (по дням/парам)
    - с метриками

Формат MVP: xlsx через openpyxl
pip install openpyxl
"""

from __future__ import annotations

from pathlib import Path
from typing import Optional
from datetime import datetime

from openpyxl import Workbook


class ExcelExportService:

    def __init__(
        self,
        teachers_repo,
        subjects_repo,
        groups_repo,
        rooms_repo,
        calendar_repo,
        curriculum_repo,
        schedule_repo,
        default_export_dir: Optional[Path] = None,
    ):
        self._teachers_repo = teachers_repo
        self._subjects_repo = subjects_repo
        self._groups_repo = groups_repo
        self._rooms_repo = rooms_repo
        self._calendar_repo = calendar_repo
        self._curriculum_repo = curriculum_repo
        self._schedule_repo = schedule_repo
        self._default_export_dir = default_export_dir or Path.cwd() / "exports"
        self._default_export_dir.mkdir(parents=True, exist_ok=True)

    # ---------------------------------------------------------
    # Public API
    # ---------------------------------------------------------

    def get_default_export_dir(self) -> Path:
        return self._default_export_dir

    # ------------------- Dictionaries ------------------------

    def export_dictionaries(self, out_dir: Path, fmt: str = "xlsx") -> Path:
        out_dir.mkdir(parents=True, exist_ok=True)

        filename = out_dir / f"dictionaries_{self._timestamp()}.xlsx"
        wb = Workbook()

        # Teachers
        ws = wb.active
        ws.title = "Teachers"
        ws.append(["ID", "Full Name", "Hard Max", "Soft Max", "Needs Method Day"])
        for t in self._teachers_repo.list_all():
            ws.append([
                t.id_teacher,
                t.full_name,
                t.hard_max_pairs_per_day,
                t.soft_max_pairs_per_day,
                t.needs_method_day,
            ])

        # Subjects
        ws = wb.create_sheet("Subjects")
        ws.append(["ID", "Name"])
        for s in self._subjects_repo.list_all():
            ws.append([s.id_subject, s.subject_name])

        # Groups
        ws = wb.create_sheet("Groups")
        ws.append(["ID", "Name", "Year", "Quantity"])
        for g in self._groups_repo.list_all():
            ws.append([g.id_group, g.group_name, g.year, g.quantity])

        # Rooms
        ws = wb.create_sheet("Rooms")
        ws.append(["ID", "Number", "Type", "Capacity", "Building"])
        for r in self._rooms_repo.list_all():
            ws.append([r.id_room, r.room_number, r.room_type, r.capacity, r.building])

        wb.save(filename)
        return filename

    # ------------------- Schedule ----------------------------

    def export_schedule_variant(
        self,
        variant_id: int,
        out_dir: Path,
        fmt: str = "xlsx",
        include_metrics: bool = True,
        include_grid_view: bool = True,
        include_table_view: bool = True,
    ) -> Path:
        out_dir.mkdir(parents=True, exist_ok=True)

        variant_dto = self._schedule_repo.get_variant_dto(variant_id)

        filename = out_dir / f"schedule_variant_{variant_id}_{self._timestamp()}.xlsx"
        wb = Workbook()

        if include_table_view:
            self._write_table_view(wb, variant_dto)

        if include_grid_view:
            self._write_grid_view(wb, variant_dto)

        if include_metrics:
            self._write_metrics(wb, variant_dto)

        wb.save(filename)
        return filename

    # ---------------------------------------------------------
    # Internal writers
    # ---------------------------------------------------------

    def _write_table_view(self, wb: Workbook, variant_dto):
        ws = wb.active
        ws.title = "Schedule Table"

        ws.append([
            "Week",
            "Week Type",
            "Day",
            "Pair",
            "Group",
            "Subject",
            "Part",
            "Teacher",
            "Room",
        ])

        for e in variant_dto.entries:
            ws.append([
                e.week_number,
                e.week_type,
                e.day_of_week,
                e.pair_number,
                e.group_name,
                e.subject_name,
                e.part_type,
                e.teacher_name,
                e.room_number,
            ])

    def _write_grid_view(self, wb: Workbook, variant_dto):
        ws = wb.create_sheet("Schedule Grid")

        # Группируем по дням и парам
        grid = {}
        for e in variant_dto.entries:
            key = (e.day_of_week, e.pair_number)
            grid.setdefault(key, []).append(
                f"{e.group_name}\n{e.subject_name}\n{e.teacher_name}\n{e.room_number}"
            )

        # Header
        ws.append(["Pair/Day", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat"])

        max_pair = max((e.pair_number for e in variant_dto.entries), default=0)

        for pair in range(1, max_pair + 1):
            row = [f"Pair {pair}"]
            for day in range(1, 7):
                cell = "\n---\n".join(grid.get((day, pair), []))
                row.append(cell)
            ws.append(row)

    def _write_metrics(self, wb: Workbook, variant_dto):
        ws = wb.create_sheet("Metrics")

        ws.append(["Variant Name", variant_dto.name])
        ws.append(["Objective Score", variant_dto.objective_score])
        ws.append([])
        ws.append(["Note", "Detailed metrics can be added here via domain.scoring"])

    # ---------------------------------------------------------

    def _timestamp(self) -> str:
        return datetime.now().strftime("%Y%m%d_%H%M%S")